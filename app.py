import threading
import logging
from datetime import datetime
from typing import Optional
from tkinter import Tk, Button, Label
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os

# Configuration constants
EXCEL_FILE_PATH = "dados_temperatura.xlsx"
CHROMEDRIVER_PATH = "C:/WebDrivers/chromedriver.exe"
HEADLESS_MODE = False  # Set to True to run Chrome in headless mode

# Setup logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def fetch_weather_data(status_label: Label) -> None:
    """
    Fetch temperature and humidity data from AccuWeather website and save to Excel file.
    Updates the status_label with progress and results.
    """
    status_label.config(text="Buscando dados...")

    options = Options()
    if HEADLESS_MODE:
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")

    service = Service(executable_path=CHROMEDRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=options)

    try:
        driver.get("https://www.accuweather.com/pt/br/s%C3%A3o-paulo/45881/current-weather/45881")

        # Wait for temperature element
        temperatura_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.current-weather-card div.temp"))
        )
        temperatura = temperatura_element.text.strip()

        # Wait for humidity element label "Umidade" and get the sibling value
        humidity_label = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Umidade')]"))
        )
        humidity_value = humidity_label.find_element(By.XPATH, "./following-sibling::div").text.strip()

        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        temperatura_formatted = f"{temperatura}"
        humidity_formatted = f"{humidity_value}"

        if not os.path.exists(EXCEL_FILE_PATH):
            wb = Workbook()
            ws = wb.active
            ws.append(["Data/Hora", "Temperatura", "Umidade"])
        else:
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active

        ws.append([data_hora, temperatura_formatted, humidity_formatted])
        wb.save(EXCEL_FILE_PATH)

        status_label.config(text=f"Capturado: {temperatura_formatted}, {humidity_formatted}")
        logging.info(f"Data captured: {temperatura_formatted}, {humidity_formatted} at {data_hora}")

    except Exception as e:
        status_label.config(text="Erro ao captar dados.")
        logging.error("Erro ao captar dados.", exc_info=True)
    finally:
        driver.quit()

def on_fetch_button_click(status_label: Label) -> None:
    """
    Run fetch_weather_data in a separate thread to avoid blocking the UI.
    """
    threading.Thread(target=fetch_weather_data, args=(status_label,), daemon=True).start()

def main() -> None:
    """
    Setup and run the Tkinter GUI application.
    """
    app = Tk()
    app.title("Captura de Temperatura SP (AccuWeather)")
    app.geometry("300x180")

    titulo = Label(app, text="Temperatura São Paulo", font=("Arial", 14))
    titulo.pack(pady=10)

    status_label = Label(app, text="", font=("Arial", 10))
    status_label.pack()

    botao = Button(
        app,
        text="Buscar previsão",
        command=lambda: on_fetch_button_click(status_label),
        width=20,
        height=2,
        bg="lightblue"
    )
    botao.pack(pady=10)

    app.mainloop()

if __name__ == "__main__":
    main()
